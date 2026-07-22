"""
생산계획 Excel -> 물류 대시보드 Push 스크립트
"""
import os
import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import sys
import re
import glob
import json
import datetime
import urllib.request
import openpyxl

# --설정 --
DASHBOARD_URL = "https://logistics-dashboard.ai-during-smart.workers.dev"
API_KEY = "logistics-api-key-2026"
PLAN_DIR = r"C:\Users\user\Desktop\생산관리\생산계획"

# material_shortage_push.py와 동일한 탐색 경로
PLAN_SEARCH_DIRS = [
    os.path.expanduser("~/Documents/HalIlApp/downloads"),
    os.path.expanduser("~/Downloads"),
    os.path.expanduser("~/Desktop/생산관리"),
    os.path.expanduser("~/Desktop"),
]

TEAMS = ["권선", "사출", "전장"]
RESIN_SPEC_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "..", "BOM_Project", "resin_spec.xlsx")

# 팀별 엑셀 구조 설정
TEAM_CONFIG = {
    "권선": {
        "plan_sheet": "권선생산계획",
        "cover_sheet": "표지",
        "date_header_row": 4,       # 날짜가 있는 행
        "sub_header_row": 5,        # 주/야 서브헤더 행
        "data_start_row": 6,        # 데이터 시작 행
        "line_col": "B",            # 호기(라인) 컬럼
        "equip_col": "B",           # 호기에서 설비명 추출
        "product_code_col": "E",    # 품번 컬럼
        "product_col": "F",         # 품명 컬럼
        "date_start_col_idx": 14,   # N열 = 14 (1-based), 첫 날짜 시작 컬럼
        "date_col_step": 4,         # 날짜당 4개 컬럼 (주/야/잔량/일소)
        "line_suffix": " line",     # 설비명 뒤에 " line" 붙임
    },
    "사출": {
        "plan_sheet": "사출생산계획",
        "cover_sheet": "표지",
        "date_header_row": 3,
        "sub_header_row": 4,
        "data_start_row": 5,
        "line_col": "C",            # 신(호기) 기준 그룹핑
        "equip_col": None,
        "line_format": "호기",       # 숫자 뒤에 "호기" 붙임
        "product_code_col": "F",    # 품번 컬럼
        "product_col": "G",         # G열이 품명 (H열은 C/T)
        "date_start_col_idx": 16,   # P열 = 16
        "date_col_step": 4,
    },
    "전장": {
        "plan_sheet": "전장생산계획",
        "cover_sheet": "표지",       # 공백 포함 가능
        "date_header_row": 5,
        "sub_header_row": 6,
        "data_start_row": 7,
        "line_col": "C",
        "equip_col": None,           # 전장은 LINE명이 곧 설비
        "product_code_col": "G",    # 품번 컬럼
        "product_col": "H",
        "date_start_col_idx": 16,   # P열 = 16
        "date_col_step": 4,         # 전장은 주간/야간/잔량/일소 = step 4 (but some have step 2)
        "line_suffix": " line",     # 설비명 뒤에 " line" 붙임
    },
}

# 전장은 2열씩 건너뛰는 구조 (주간+잔량만)인지 확인 필요 → 동적 감지


def load_resin_spec():
    """resin_spec.xlsx + BOM 기반 품번→수지 매핑 구축.

    Returns:
        dict: {제품품번: {resin, grade, temp, time, raw_code}} 또는 빈 dict
    """
    spec_path = os.path.normpath(RESIN_SPEC_PATH)
    if not os.path.isfile(spec_path):
        return {}

    # 1. resin_spec.xlsx → 원자재 품번→수지 정보
    try:
        wb = openpyxl.load_workbook(spec_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        resin_by_raw = {}  # 원자재품번 → {resin, grade, temp, time}
        cur_resin = cur_grade = cur_temp = cur_time = ""
        for r in range(2, ws.max_row + 1):
            no_val = ws.cell(row=r, column=1).value
            resin = ws.cell(row=r, column=2).value
            code = ws.cell(row=r, column=3).value
            grade = ws.cell(row=r, column=4).value
            temp = ws.cell(row=r, column=6).value
            time_val = ws.cell(row=r, column=7).value
            # 새 수지 항목(NO. 있음) 시작 시 온도/시간 리셋
            if no_val is not None and resin:
                cur_resin = str(resin).strip()
                cur_grade = str(grade).strip() if grade else ""
                cur_temp = str(temp).strip() if temp else ""
                cur_time = str(time_val).strip() if time_val else ""
            else:
                if resin: cur_resin = str(resin).strip()
                if grade: cur_grade = str(grade).strip()
                if temp: cur_temp = str(temp).strip()
                if time_val: cur_time = str(time_val).strip()
            if code:
                raw_code = str(code).strip()
                if raw_code not in resin_by_raw:
                    resin_by_raw[raw_code] = {
                        "resin": cur_resin, "grade": cur_grade,
                        "temp": cur_temp, "time": cur_time,
                    }
        wb.close()
    except Exception as e:
        print(f"  [WARN] resin_spec 로드 실패: {e}")
        return {}

    # 2. BOM 로드 → 제품 품번→수지 원자재 매핑
    bom_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "..", "BOM_Project", "BOM_DATA.xlsx")
    if not os.path.isfile(bom_path):
        print(f"  [WARN] BOM 파일 없음: {bom_path}")
        return {}

    try:
        import pandas as pd
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                        "..", "BOM_Project"))
        from bom import normalize_bom
        bom_raw = pd.read_excel(bom_path)
        bom_df = normalize_bom(bom_raw)
    except Exception as e:
        print(f"  [WARN] BOM 로드 실패: {e}")
        return {}

    # BOM 그래프: 모품목 → [자재품번]
    children_map = {}
    for _, row in bom_df.iterrows():
        parent = str(row["모품목"]).strip()
        child = str(row["자재품번"]).strip()
        if parent not in children_map:
            children_map[parent] = []
        children_map[parent].append(child)

    resin_raw_codes = set(resin_by_raw.keys())

    # BFS로 제품→수지 원자재 탐색 (최대 4레벨)
    product_resin = {}  # 제품품번 → {resin, grade, temp, time, raw_code}

    def _find_resin(product_code, max_depth=4):
        """BOM을 따라 내려가며 수지 원자재를 찾음"""
        visited = set()
        queue = [(product_code, 0)]
        while queue:
            node, depth = queue.pop(0)
            if node in visited or depth > max_depth:
                continue
            visited.add(node)
            for child in children_map.get(node, []):
                if child in resin_raw_codes:
                    return child
                queue.append((child, depth + 1))
        return None

    # 모든 BOM 루트 품목에 대해 수지 매핑 구축
    all_parents = set(children_map.keys())
    for parent_code in all_parents:
        raw = _find_resin(parent_code)
        if raw:
            info = resin_by_raw[raw]
            product_resin[parent_code] = {
                **info, "raw_code": raw,
            }

    print(f"  수지 스펙: 원자재 {len(resin_by_raw)}종 → 제품 {len(product_resin)}건 매핑")
    return product_resin


def detect_resin_changes(switches, resin_spec):
    """ITEM 변경 시 수지가 달라지는 경우 감지 (품번 기준)"""
    if not resin_spec:
        return []
    changes = []
    for sw in switches:
        from_code = sw.get("fromCode", "")
        to_code = sw.get("toCode", "")
        from_info = resin_spec.get(from_code)
        to_info = resin_spec.get(to_code)
        if from_info and to_info and from_info["resin"] != to_info["resin"]:
            changes.append({
                "line": sw["line"],
                "equip": sw.get("equip", ""),
                "shift": sw.get("shift", ""),
                "from": sw["from"],
                "to": sw["to"],
                "resin": to_info["resin"],
                "resinGrade": to_info.get("grade", ""),
                "fromResin": from_info["resin"],
                "fromGrade": from_info.get("grade", ""),
                "temp": to_info["temp"],
                "time": to_info["time"],
            })
    return changes


def detect_resin_for_restarts(resume_details, resin_spec):
    """재가동 라인의 수지 사전 준비 정보 생성 (후공정 라인 제외됨)"""
    if not resin_spec or not resume_details:
        return []
    preps = []
    for rd in resume_details:
        code = rd.get("code", "")
        info = resin_spec.get(code)
        if info:
            preps.append({
                "line": rd["line"],
                "equip": rd.get("equip", ""),
                "shift": rd.get("shift", ""),
                "from": "(재가동)",
                "to": rd["product"],
                "resin": info["resin"],
                "resinGrade": info.get("grade", ""),
                "fromResin": "",
                "fromGrade": "",
                "temp": info["temp"],
                "time": info["time"],
                "type": "restart",
            })
    return preps


def find_latest_folder(base_dir):
    """가장 최근 폴더를 찾음 (예: '7월17일')"""
    if not os.path.isdir(base_dir):
        return None
    folders = []
    for name in os.listdir(base_dir):
        path = os.path.join(base_dir, name)
        if os.path.isdir(path):
            folders.append((os.path.getmtime(path), name, path))
    if not folders:
        return None
    folders.sort(reverse=True)
    return folders[0][2]


def find_team_file(folder, team):
    """폴더에서 팀명이 포함된 Excel 파일 찾기"""
    if not folder:
        return None
    for ext in ("*.xlsx", "*.xlsm"):
        for f in glob.glob(os.path.join(folder, ext)):
            basename = os.path.basename(f)
            if team in basename and "~$" not in basename:
                return f
    return None


def find_team_file_broad(team):
    """여러 경로에서 팀 생산계획 파일 탐색 (HalIlApp/downloads 우선)"""
    pattern = f"{team}*생산계획*"
    candidates = []
    # 1. GW 다운로드 폴더 우선 탐색
    for search_dir in PLAN_SEARCH_DIRS:
        if not os.path.isdir(search_dir):
            continue
        for f in glob.glob(os.path.join(search_dir, pattern + ".xlsx")) + glob.glob(os.path.join(search_dir, pattern + ".xlsm")):
            if "~$" not in f:
                candidates.append(f)
    # 2. 생산계획 폴더 (하위 폴더 포함)
    if os.path.isdir(PLAN_DIR):
        for root, dirs, files in os.walk(PLAN_DIR):
            for f in glob.glob(os.path.join(root, pattern + ".xlsx")) + glob.glob(os.path.join(root, pattern + ".xlsm")):
                if "~$" not in f:
                    candidates.append(f)
    if not candidates:
        return None
    # 최신 파일 반환
    candidates.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return candidates[0]


def col_letter_to_idx(letter):
    """A=1, B=2, ..., Z=26, AA=27"""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def find_date_column(ws, header_row, target_date, start_col=1, max_col=None):
    """헤더 행에서 target_date에 해당하는 컬럼 인덱스 찾기"""
    if max_col is None:
        max_col = ws.max_column
    for col in range(start_col, min(max_col + 1, 250)):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, datetime.datetime):
            if val.strftime("%Y-%m-%d") == target_date:
                return col
    return None


def detect_date_step(ws, header_row, first_date_col):
    """날짜 컬럼 간격 자동 감지"""
    first_val = ws.cell(row=header_row, column=first_date_col).value
    if not isinstance(first_val, datetime.datetime):
        return 4  # default
    # 다음 날짜 컬럼 찾기
    for step in range(2, 8):
        next_val = ws.cell(row=header_row, column=first_date_col + step).value
        if isinstance(next_val, datetime.datetime):
            return step
    return 4


def parse_cover_notices(wb, cover_sheet_name):
    """표지 시트에서 특이사항 추출"""
    notices = []
    # 표지 시트 찾기 (공백 포함)
    ws = None
    for sn in wb.sheetnames:
        if "표지" in sn:
            ws = wb[sn]
            break
    if not ws:
        return notices

    # B열=No, C열=호기(tag), D열=내용(text) 패턴
    for row in range(7, min(ws.max_row + 1, 35)):
        tag_val = ws.cell(row=row, column=3).value  # C열 = 호기
        text_val = ws.cell(row=row, column=4).value  # D열 = 내용
        if text_val and str(text_val).strip():
            text = str(text_val).strip()
            # "■ " 접두사 제거
            if text.startswith("■"):
                text = text[1:].strip()
            tag = str(tag_val).strip() if tag_val else "공지"
            notices.append({"tag": tag, "text": text})
    return notices


def parse_plan_sheet(wb, team, config, target_date, prev_date, next_date=None):
    """생산계획 시트에서 해당 날짜의 생산 데이터 추출"""
    ws = None
    for sn in wb.sheetnames:
        if config["plan_sheet"] in sn:
            ws = wb[sn]
            break
    if not ws:
        print(f"  [!] {config['plan_sheet']} 시트를 찾을 수 없음")
        return None

    header_row = config["date_header_row"]

    # 날짜 컬럼 찾기
    target_col = find_date_column(ws, header_row, target_date)
    if not target_col:
        print(f"  [!] {target_date} 날짜 컬럼을 찾을 수 없음")
        return None

    # 날짜 간격 자동 감지
    date_step = detect_date_step(ws, header_row, target_col)
    print(f"  날짜 컬럼: {target_col}, 간격: {date_step}")

    # 전날 컬럼
    prev_col = find_date_column(ws, header_row, prev_date) if prev_date else None

    # 명일 컬럼
    next_col = find_date_column(ws, header_row, next_date) if next_date else None

    # 서브헤더 확인 (주/야 or 주간/야간)
    # target_col = 주간, target_col+1 = 야간
    day_col = target_col
    night_col = target_col + 1

    line_col_idx = col_letter_to_idx(config["line_col"])
    product_col_idx = col_letter_to_idx(config["product_col"])
    product_code_col_idx = col_letter_to_idx(config["product_code_col"]) if config.get("product_code_col") else None
    equip_col_idx = col_letter_to_idx(config["equip_col"]) if config.get("equip_col") else None
    data_start = config["data_start_row"]

    # 데이터 추출
    current_line = ""
    current_equip = ""
    line_equip_map = {}        # line -> 설비명
    lines_day = set()
    lines_night = set()
    switches = []
    line_products_today = {}   # line -> [(product_code, product_name, day_qty, night_qty)]
    line_products_prev = {}    # line -> [(product_code, product_name, day_qty, night_qty)]
    line_products_next = {}    # line -> [(product_code, product_name, next_day_qty)]

    for row in range(data_start, ws.max_row + 1):
        # 라인 번호 (비어있으면 이전 값 유지, 개행 제거)
        line_val = ws.cell(row=row, column=line_col_idx).value
        if line_val and str(line_val).strip():
            current_line = re.sub(r'\s+', ' ', str(line_val)).strip()
        # 설비명
        if equip_col_idx:
            equip_val = ws.cell(row=row, column=equip_col_idx).value
            if equip_val and str(equip_val).strip():
                raw_equip = re.sub(r'\s+', ' ', str(equip_val)).strip()
                # "호" 포맷 적용 (사출: "4" → "4호")
                equip_fmt = config.get("equip_format", "")
                if equip_fmt and raw_equip.replace("-", "").replace(".", "").isdigit():
                    current_equip = f"{raw_equip}{equip_fmt}"
                else:
                    current_equip = raw_equip
            if current_line and current_equip:
                line_equip_map[current_line] = current_equip

        if not current_line:
            continue

        # 품번 (품명보다 먼저 읽음)
        product_code = ""
        if product_code_col_idx:
            code_val = ws.cell(row=row, column=product_code_col_idx).value
            if code_val:
                product_code = str(code_val).strip().replace(".0", "")

        # 품명 (개행, 특수공백 제거)
        product_val = ws.cell(row=row, column=product_col_idx).value
        if not product_val:
            continue
        product = re.sub(r'\s+', ' ', str(product_val)).strip()
        if not product:
            continue

        # 오늘 생산량
        day_qty = ws.cell(row=row, column=day_col).value or 0
        night_qty = ws.cell(row=row, column=night_col).value or 0
        try:
            day_qty = float(day_qty) if day_qty else 0
            night_qty = float(night_qty) if night_qty else 0
        except (ValueError, TypeError):
            day_qty = 0
            night_qty = 0

        if day_qty > 0 or night_qty > 0:
            if current_line not in line_products_today:
                line_products_today[current_line] = []
            line_products_today[current_line].append((product_code, product, day_qty, night_qty))
            if day_qty > 0:
                lines_day.add(current_line)
            if night_qty > 0:
                lines_night.add(current_line)

        # 전날 생산량 (스위치 감지용)
        if prev_col:
            prev_day_qty = ws.cell(row=row, column=prev_col).value or 0
            prev_night_qty = ws.cell(row=row, column=prev_col + 1).value or 0
            try:
                prev_day_qty = float(prev_day_qty) if prev_day_qty else 0
                prev_night_qty = float(prev_night_qty) if prev_night_qty else 0
            except (ValueError, TypeError):
                prev_day_qty = 0
                prev_night_qty = 0

            if prev_day_qty > 0 or prev_night_qty > 0:
                if current_line not in line_products_prev:
                    line_products_prev[current_line] = []
                line_products_prev[current_line].append((product_code, product, prev_day_qty, prev_night_qty))

        # 명일 주간 생산량
        if next_col:
            next_day_qty = ws.cell(row=row, column=next_col).value or 0
            try:
                next_day_qty = float(next_day_qty) if next_day_qty else 0
            except (ValueError, TypeError):
                next_day_qty = 0

            if next_day_qty > 0:
                if current_line not in line_products_next:
                    line_products_next[current_line] = []
                line_products_next[current_line].append((product_code, product, next_day_qty))

    # 출력용: equip를 line으로 표시할지 여부
    use_equip_as_line = config.get("display_equip_as_line", False)

    # ITEM 변경 감지 제외 패턴: 후공정/이관/조립 등 복합 라인
    _SKIP_SWITCH_PATTERNS = {"이관", "융착", "조립", "오링", "OS", "WC-PSV"}

    def _is_multi_process_line(line_name):
        """복수 공정이 혼재된 라인인지 판별 (ITEM 변경 비교 무의미)"""
        return any(kw in line_name for kw in _SKIP_SWITCH_PATTERNS)

    # --- 시프트별 메인 품목 헬퍼 ---
    def _shift_main(prods, shift):
        """shift='day'→주간, 'night'→야간 중 최대수량 품목. (code, name, qty) or None"""
        if shift == 'day':
            items = [(c, n, d) for c, n, d, ng in prods if d > 0]
        else:
            items = [(c, n, ng) for c, n, d, ng in prods if ng > 0]
        if not items:
            return None
        items.sort(key=lambda x: x[2], reverse=True)
        return items[0]

    def _shift_keys(prods, shift):
        """시프트 내 모든 품목 키 set"""
        if shift == 'day':
            return set((c or n) for c, n, d, ng in prods if d > 0)
        else:
            return set((c or n) for c, n, d, ng in prods if ng > 0)

    def _format_line(line_name):
        """호기/line 표시"""
        line_fmt = config.get("line_format", "")
        line_sfx = config.get("line_suffix", "")
        if line_fmt and re.match(r'^[\d/\-\.]+$', line_name):
            return f"{line_name}{line_fmt}"
        elif line_sfx:
            return f"{line_name}{line_sfx}"
        return line_name

    def _is_post_process(line_name):
        """후공정 라인 판별 (사출: 비숫자 설비명 포함)"""
        if _is_multi_process_line(line_name) or "후공정" in line_name:
            return True
        line_fmt = config.get("line_format", "")
        if line_fmt and not re.match(r'^[\d/\-\.]+$', line_name):
            return True
        return False

    # --- 시프트 내 품목 리스트 헬퍼 ---
    _PAREN_CODE_RE = re.compile(r'\((\d{7,})\)')

    def _product_key(code, name):
        """품목 동일성 판별 키: 품명 내 괄호 관리번호 우선, 없으면 품번, 최후 품명"""
        m = _PAREN_CODE_RE.search(name or "")
        if m:
            return m.group(1)
        return code or name or ""

    def _shift_products(prods, shift):
        """시프트 내 품목을 Excel 행 순서대로 반환 (관리번호/품번 기준 중복 제거). [(code, name, qty)]"""
        seen = set()
        result = []
        for c, n, d, ng in prods:
            qty = d if shift == 'day' else ng
            if qty > 0:
                key = _product_key(c, n)
                if key not in seen:
                    seen.add(key)
                    result.append((c, n, qty))
        return result

    # --- ITEM 변경 감지 (시프트별) ---
    # 튜플 구조: (product_code, product_name, day_qty, night_qty)
    for line in line_products_today:
        if _is_multi_process_line(line):
            continue
        today_prods = line_products_today[line]
        prev_prods = line_products_prev.get(line)

        equip_val = line_equip_map.get(line, "")
        if use_equip_as_line and equip_val:
            display_line = _format_line(equip_val)
            display_equip = ""
        else:
            display_line = _format_line(line)
            display_equip = equip_val

        # (1) 주간 ITEM변경: 전날 전체(주+야 합산) 메인품목 → 오늘 주간
        today_day = _shift_main(today_prods, 'day')
        if today_day and prev_prods:
            # 전날 전체 합산 기준 메인품목
            prev_sorted = sorted(prev_prods, key=lambda x: x[2] + x[3], reverse=True)
            prev_code = prev_sorted[0][0]
            prev_name = prev_sorted[0][1]
            prev_total = int(prev_sorted[0][2] + prev_sorted[0][3])
            today_key = today_day[0] or today_day[1]
            prev_key = prev_code or prev_name
            if today_key != prev_key:
                prev_all_keys = set((p[0] or p[1]) for p in prev_prods)
                today_day_keys = _shift_keys(today_prods, 'day')
                if not (today_key in prev_all_keys and prev_key in today_day_keys):
                    switches.append({
                        "line": display_line, "equip": display_equip,
                        "shift": "주간",
                        "fromCode": prev_code, "toCode": today_day[0],
                        "from": prev_name, "to": today_day[1],
                        "qty": int(today_day[2]), "fromQty": prev_total,
                    })

        # (1-b) 주간 중 ITEM변경: 같은 주간 시프트 내 복수 품목 (첫→끝 압축)
        day_items = _shift_products(today_prods, 'day')
        # 같은 품번끼리만 다른 건 제외 (공정명만 다른 경우)
        day_items = [it for i, it in enumerate(day_items)
                     if i == 0 or (it[0] != day_items[i-1][0] or not it[0])]
        if len(day_items) >= 2:
            mid_names = [it[1] for it in day_items[1:-1]]
            sw_entry = {
                "line": display_line, "equip": display_equip,
                "shift": "주간 중",
                "fromCode": day_items[0][0], "toCode": day_items[-1][0],
                "from": day_items[0][1], "to": day_items[-1][1],
                "qty": int(day_items[-1][2]), "fromQty": int(day_items[0][2]),
            }
            if mid_names:
                sw_entry["midCount"] = len(day_items) - 1
            switches.append(sw_entry)

        # (2) 교대 후 ITEM변경: 오늘 주간 → 오늘 야간
        today_night = _shift_main(today_prods, 'night')
        if today_day and today_night:
            day_key = today_day[0] or today_day[1]
            night_key = today_night[0] or today_night[1]
            if day_key != night_key:
                day_ks = _shift_keys(today_prods, 'day')
                night_ks = _shift_keys(today_prods, 'night')
                if not (night_key in day_ks and day_key in night_ks):
                    switches.append({
                        "line": display_line, "equip": display_equip,
                        "shift": "교대 후",
                        "fromCode": today_day[0], "toCode": today_night[0],
                        "from": today_day[1], "to": today_night[1],
                        "qty": int(today_night[2]), "fromQty": int(today_day[2]),
                    })

        # (2-b) 야간 중 ITEM변경: 같은 야간 시프트 내 복수 품목 (첫→끝 압축)
        night_items = _shift_products(today_prods, 'night')
        # 같은 품번끼리만 다른 건 제외
        night_items = [it for i, it in enumerate(night_items)
                       if i == 0 or (it[0] != night_items[i-1][0] or not it[0])]
        if len(night_items) >= 2:
            mid_names = [it[1] for it in night_items[1:-1]]
            sw_entry = {
                "line": display_line, "equip": display_equip,
                "shift": "야간 중",
                "fromCode": night_items[0][0], "toCode": night_items[-1][0],
                "from": night_items[0][1], "to": night_items[-1][1],
                "qty": int(night_items[-1][2]), "fromQty": int(night_items[0][2]),
            }
            if mid_names:
                sw_entry["midCount"] = len(night_items) - 1
            switches.append(sw_entry)

        # (3) 명일 주간 ITEM변경: 금일 야간 → 명일 주간
        next_prods = line_products_next.get(line)
        if next_prods and today_night:
            next_sorted = sorted(next_prods, key=lambda x: x[2], reverse=True)
            next_main = next_sorted[0]  # (code, name, qty)
            night_key = today_night[0] or today_night[1]
            next_key = next_main[0] or next_main[1]
            if night_key != next_key:
                night_ks = _shift_keys(today_prods, 'night')
                next_ks = set((c or n) for c, n, q in next_prods)
                if not (next_key in night_ks and night_key in next_ks):
                    switches.append({
                        "line": display_line, "equip": display_equip,
                        "shift": "명일 주간",
                        "fromCode": today_night[0], "toCode": next_main[0],
                        "from": today_night[1], "to": next_main[1],
                        "qty": int(next_main[2]), "fromQty": int(today_night[2]),
                    })
        elif next_prods and not today_night and today_day:
            # 금일 야간 미가동 시 → 금일 주간 기준으로 비교
            next_sorted = sorted(next_prods, key=lambda x: x[2], reverse=True)
            next_main = next_sorted[0]
            day_key = today_day[0] or today_day[1]
            next_key = next_main[0] or next_main[1]
            if day_key != next_key:
                day_ks = _shift_keys(today_prods, 'day')
                next_ks = set((c or n) for c, n, q in next_prods)
                if not (next_key in day_ks and day_key in next_ks):
                    switches.append({
                        "line": display_line, "equip": display_equip,
                        "shift": "명일 주간",
                        "fromCode": today_day[0], "toCode": next_main[0],
                        "from": today_day[1], "to": next_main[1],
                        "qty": int(next_main[2]), "fromQty": int(today_day[2]),
                    })

    # --- 재가동 감지 (전날 생산 없고 오늘 생산 있는 라인, 시프트 표시) ---
    resumes = []
    resume_details = []  # 수지 사전 준비 감지용 (후공정 제외)
    for line in line_products_today:
        if line not in line_products_prev:
            today_prods = line_products_today[line]
            today_day = _shift_main(today_prods, 'day')
            today_night = _shift_main(today_prods, 'night')

            equip_val = line_equip_map.get(line, "")
            display = equip_val if use_equip_as_line and equip_val else line
            display_fmt = _format_line(display)

            # 주간부터 가동이면 주간, 야간만이면 야간
            shift = "주간" if today_day else "야간"
            resumes.append(f"{display_fmt}({shift})")

            # 후공정 제외하고 수지 사전 준비용 상세 정보
            if not _is_post_process(line):
                main_prod = today_day or today_night
                resume_details.append({
                    "line": display_fmt,
                    "shift": shift,
                    "equip": equip_val if not use_equip_as_line else "",
                    "code": main_prod[0],
                    "product": main_prod[1],
                })

    # 명일 주간 재가동 (금일 전체 미가동 → 명일 주간 가동)
    for line in line_products_next:
        if line not in line_products_today and line not in line_products_prev:
            next_prods = line_products_next[line]
            next_main = sorted(next_prods, key=lambda x: x[2], reverse=True)[0]

            equip_val = line_equip_map.get(line, "")
            display = equip_val if use_equip_as_line and equip_val else line
            display_fmt = _format_line(display)

            resumes.append(f"{display_fmt}(명일 주간)")

            if not _is_post_process(line):
                resume_details.append({
                    "line": display_fmt,
                    "shift": "명일 주간",
                    "equip": equip_val if not use_equip_as_line else "",
                    "code": next_main[0],
                    "product": next_main[1],
                })

    return {
        "lines": {"day": len(lines_day), "night": len(lines_night)},
        "switches": switches,
        "resumes": resumes,
        "resumeDetails": resume_details,
        "lineProducts": {
            line: [{"code": c, "product": p, "dayQty": int(d), "nightQty": int(n)}
                   for c, p, d, n in prods]
            for line, prods in line_products_today.items()
        },
    }


def parse_team(folder, team, target_date, prev_date, next_date=None):
    """팀별 Excel 파일 파싱 (GW 다운로드 최신 파일 우선)"""
    # GW 다운로드 등 전체 경로에서 최신 파일 우선 탐색
    filepath = find_team_file_broad(team)
    # 못 찾으면 지정 폴더에서 검색
    if not filepath and folder:
        filepath = find_team_file(folder, team)
    if not filepath:
        print(f"  [!] {team} 파일 없음")
        return None

    print(f"  파일: {os.path.basename(filepath)}")
    config = TEAM_CONFIG[team]

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
    except Exception as e:
        print(f"  [!] 파일 열기 실패: {e}")
        return None

    # 표지 → 특이사항
    notices = parse_cover_notices(wb, config.get("cover_sheet", "표지"))
    print(f"  특이사항: {len(notices)}건")

    # 생산계획 → 라인/스위치/재가동
    plan_data = parse_plan_sheet(wb, team, config, target_date, prev_date, next_date)
    if not plan_data:
        return {"notices": notices, "lines": {"day": 0, "night": 0},
                "switches": [], "resumes": []}

    plan_data["notices"] = notices
    print(f"  라인: 주간 {plan_data['lines']['day']}개, 야간 {plan_data['lines']['night']}개")
    sw_day = [s for s in plan_data['switches'] if s.get('shift') == '주간']
    sw_handover = [s for s in plan_data['switches'] if s.get('shift') == '교대 후']
    sw_next = [s for s in plan_data['switches'] if s.get('shift') == '명일 주간']
    print(f"  ITEM변경: 주간 {len(sw_day)}건, 교대 후 {len(sw_handover)}건, 명일 주간 {len(sw_next)}건")
    print(f"  재가동: {len(plan_data['resumes'])}개")
    return plan_data


def push_to_dashboard(payload, target_date):
    """대시보드 API로 push"""
    url = f"{DASHBOARD_URL}/api/plan/push"
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "X-API-Key": API_KEY,
            "User-Agent": "ProductionPlanPush/1.0",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode("utf-8"))
            print(f"\n[OK] Push 성공: {result}")
            return True
    except Exception as e:
        print(f"\n[ERROR] Push 실패: {e}")
        return False


PUSH_STAMP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               ".last_push.json")


def load_push_stamps():
    """팀별 마지막 push 시 파일 수정시간 기록 로드"""
    if os.path.isfile(PUSH_STAMP_FILE):
        try:
            with open(PUSH_STAMP_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_push_stamps(stamps):
    """팀별 파일 수정시간 기록 저장"""
    with open(PUSH_STAMP_FILE, "w", encoding="utf-8") as f:
        json.dump(stamps, f, ensure_ascii=False, indent=2)


def main():
    # --force 플래그: 수정시간 무시하고 전체 push
    force_mode = "--force" in sys.argv

    # 날짜 결정: 인자로 날짜 지정 가능, 없으면 내일
    if len(sys.argv) > 1 and len(sys.argv[1]) == 10:
        target_date = sys.argv[1]
    else:
        # 08시 기준 날짜 (KST): 08시 이전이면 전날이 기준일
        now_kst = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=9)
        if now_kst.hour < 8:
            target_date = (now_kst - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
        else:
            target_date = now_kst.strftime("%Y-%m-%d")

    # 전날 (주말이면 금요일로)
    td = datetime.datetime.strptime(target_date, "%Y-%m-%d")
    prev_td = td - datetime.timedelta(days=1)
    # 전날이 일요일이면 금요일(-2), 토요일이면 금요일(-1)
    if prev_td.weekday() == 6:  # 일요일
        prev_td = prev_td - datetime.timedelta(days=2)
    elif prev_td.weekday() == 5:  # 토요일
        prev_td = prev_td - datetime.timedelta(days=1)
    prev_date = prev_td.strftime("%Y-%m-%d")

    # 명일 (주말이면 월요일로)
    next_td = td + datetime.timedelta(days=1)
    if next_td.weekday() == 5:  # 토요일 → 월요일
        next_td = next_td + datetime.timedelta(days=2)
    elif next_td.weekday() == 6:  # 일요일 → 월요일
        next_td = next_td + datetime.timedelta(days=1)
    next_date = next_td.strftime("%Y-%m-%d")

    # 폴더 인자 또는 최신 폴더
    if len(sys.argv) > 2 and not sys.argv[2].startswith("--"):
        folder = sys.argv[2]
    else:
        folder = find_latest_folder(PLAN_DIR)

    print(f"=== 생산계획 Push ===")
    print(f"대상 날짜: {target_date}")
    print(f"전날: {prev_date}")
    print(f"명일: {next_date}")
    if folder:
        print(f"폴더: {os.path.basename(folder)}")
    else:
        print(f"폴더: 자동 탐색 모드")
    if force_mode:
        print(f"모드: --force (전체 갱신)")
    print()

    # 수지 스펙 로드
    resin_spec = load_resin_spec()
    if resin_spec:
        print(f"수지 스펙: {len(resin_spec)}개 품목 로드")
    else:
        print("수지 스펙: 미발견 (수지 변경 감지 불가)")
    print()

    # 마지막 push 기록 로드
    stamps = load_push_stamps()
    date_stamps = stamps.get(target_date, {})

    # 팀별 파일 탐색 + 수정시간 비교
    team_files = {}
    for team in TEAMS:
        filepath = find_team_file_broad(team)
        if not filepath and folder:
            filepath = find_team_file(folder, team)
        if not filepath:
            print(f"  {team}: 파일 없음 → 기존 유지")
            continue
        mtime = os.path.getmtime(filepath)
        last_mtime = date_stamps.get(team, {}).get("mtime", 0)
        if not force_mode and mtime <= last_mtime:
            print(f"  {team}: 변경 없음 ({os.path.basename(filepath)}) → 기존 유지")
            continue
        team_files[team] = {"path": filepath, "mtime": mtime}
        marker = "신규" if last_mtime == 0 else "갱신"
        print(f"  {team}: {marker} ({os.path.basename(filepath)})")
    print()

    if not team_files:
        print("[INFO] 변경된 생산계획 없음 — push 생략")
        return

    # 변경된 팀만 파싱
    teams_data = {}
    total_day = 0
    total_night = 0
    total_switches = 0
    total_resumes = 0

    for team, finfo in team_files.items():
        print(f"--{team} --")
        result = parse_team(folder, team, target_date, prev_date, next_date)
        if result:
            switches = result.get("switches", [])
            # 수지 정보는 사출만 적용 (권선/전장은 수지 건조 불필요)
            if team == "사출" and resin_spec:
                resin_changes = detect_resin_changes(switches, resin_spec)
                resin_restarts = detect_resin_for_restarts(
                    result.get("resumeDetails", []), resin_spec)
            else:
                resin_changes = []
                resin_restarts = []

            # 수지 정보를 switches에 병합 (ITEM변경 + 건조기 사전준비 통합)
            resin_by_key = {}
            for rc in resin_changes:
                resin_by_key[(rc['line'], rc.get('shift', ''))] = rc
            for sw in switches:
                key = (sw['line'], sw.get('shift', ''))
                if key in resin_by_key:
                    rc = resin_by_key.pop(key)
                    sw['resin'] = rc['resin']
                    sw['resinGrade'] = rc.get('resinGrade', '')
                    sw['fromResin'] = rc.get('fromResin', '')
                    sw['fromResinGrade'] = rc.get('fromGrade', '')
                    sw['temp'] = rc['temp']
                    sw['time'] = rc['time']
            # 매칭 안 된 수지변경 → 별도 switch 엔트리로 추가
            for rc in resin_by_key.values():
                switches.append({
                    "line": rc['line'], "equip": rc.get('equip', ''),
                    "shift": rc.get('shift', ''),
                    "from": rc.get('from', ''), "to": rc.get('to', ''),
                    "resin": rc['resin'], "resinGrade": rc.get('resinGrade', ''),
                    "fromResin": rc.get('fromResin', ''), "fromResinGrade": rc.get('fromGrade', ''),
                    "temp": rc['temp'], "time": rc['time'],
                })
            # 재가동 수지준비 → switch 엔트리로 추가
            for rc in resin_restarts:
                switches.append({
                    "line": rc['line'], "equip": rc.get('equip', ''),
                    "shift": rc.get('shift', ''),
                    "from": rc.get('from', ''), "to": rc.get('to', ''),
                    "resin": rc['resin'], "resinGrade": rc.get('resinGrade', ''),
                    "fromResin": '', "fromResinGrade": '',
                    "temp": rc['temp'], "time": rc['time'],
                    "type": "restart",
                })

            teams_data[team] = {
                "lines": result["lines"],
                "notices": result.get("notices", []),
                "switches": switches,
            }
            total_day += result["lines"]["day"]
            total_night += result["lines"]["night"]
            total_switches += len(switches)
            total_resumes += len(result.get("resumes", []))
            for sw in switches:
                shift_tag = f"[{sw.get('shift', '')}]"
                resin_tag = ""
                if sw.get('resin'):
                    fr = sw.get('fromResin', '')
                    grade = sw.get('resinGrade', '')
                    resin_name = f"{sw['resin']} {grade}" if grade else sw['resin']
                    resin_tag = f" 🔸{fr}->{resin_name}({sw['temp']}/{sw['time']})" if fr else f" 🔸{resin_name}({sw['temp']}/{sw['time']})"
                sw_type = "(재가동)" if sw.get('type') == 'restart' else ""
                print(f"  ITEM변경{shift_tag}: {sw['line']}{sw_type} {sw['from']} → {sw['to']}{resin_tag}")
        print()

    if not teams_data:
        print("[WARN] 파싱 성공한 팀 없음 — push 생략")
        return

    # 페이로드 구성 (변경된 팀만 포함 → worker에서 기존 데이터와 merge)
    now_kst = datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=9)
    payload = {
        "date": target_date,
        "teams": teams_data,
        "pushedAt": now_kst.isoformat(),
        "source": "production_plan_push.py",
    }

    pushed_teams = list(teams_data.keys())
    print(f"=== 요약 ===")
    print(f"  갱신 팀: {', '.join(pushed_teams)}")
    print(f"  주간 라인: {total_day}개")
    print(f"  야간 라인: {total_night}개")
    print(f"  ITEM 변경: {total_switches}건")
    print(f"  재가동: {total_resumes}건")
    print()

    # Push
    success = push_to_dashboard(payload, target_date)

    # 성공 시 수정시간 기록 저장
    if success:
        if target_date not in stamps:
            stamps[target_date] = {}
        for team, finfo in team_files.items():
            if team in teams_data:
                stamps[target_date][team] = {
                    "mtime": finfo["mtime"],
                    "file": os.path.basename(finfo["path"]),
                    "pushedAt": now_kst.strftime("%Y-%m-%d %H:%M:%S"),
                }
        # 오래된 날짜 기록 정리 (7일 이상 전)
        cutoff = (td - datetime.timedelta(days=7)).strftime("%Y-%m-%d")
        stamps = {k: v for k, v in stamps.items() if k >= cutoff}
        save_push_stamps(stamps)


if __name__ == "__main__":
    main()
